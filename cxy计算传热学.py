import csv
import pandas
import openpyxl

length = 3
height = 2
TL, TR, TB, TU = 3, 5, 2, 4
den, cv, k =  100, 1000, 10
s = 10
dt = 200
n, m = 5, 5
maxstep = 20
tr = 1
miter = 1
maxiter = 1000
maxresi = 1**-7
list1 = []
list2 = []
T0 = [[0 for i in range(n+1)]for j in range(m+1)]
T =  [[0 for i in range(n+2)]for j in range(m+2)]
Tn0 =[[0 for i in range(n+2)]for j in range(m+2)]
ae0= [[0 for i in range(n+1)]for j in range(m+1)]
aw0 =[[0 for i in range(n+1)]for j in range(m+1)]
an0 =[[0 for i in range(n+1)]for j in range(m+1)]
as0 =[[0 for i in range(n+1)]for j in range(m+1)]
ap0 =[[0 for i in range(n+1)]for j in range(m+1)]
ap1 =[[0 for i in range(n+1)]for j in range(m+1)]
bp = [[0 for i in range(n+1)]for j in range(m+1)]
ae1 =[[0 for i in range(n+1)]for j in range(m+1)]
aw1 =[[0 for i in range(n+1)]for j in range(m+1)]
an1 =[[0 for i in range(n+1)]for j in range(m+1)]
as1 =[[0 for i in range(n+1)]for j in range(m+1)]
resi = [[0 for i in range(n+1)]for j in range(m+1)]

dx = length/n
dy = height/m
if tr == 0:
    for i in range(2, n):
        for j in range(2, m):
            ae0[i][j] = k * dy / dx
            aw0[i][j] = k * dy / dx
            an0[i][j] = k * dx / dy
            as0[i][j] = k * dx / dy
            ap0[i][j] = den * cv * dx * dy / dt - ae0[i][j] - aw0[i][j] - an0[i][j] - as0[i][j]
            ap1[i][j] = ap0[i][j] + ae0[i][j] + aw0[i][j] + an0[i][j] + as0[i][j]
            bp[i][j] = s * dx * dy
    i = 1
    for j in range(2, m):
        ae0[i][j] = k * dy / dx
        aw0[i][j] = k * dy / (dx / 2)
        an0[i][j] = k * dx / dy
        as0[i][j] = k * dx / dy
        ap0[i][j] = den * cv * dx * dy / dt - ae0[i][j] - aw0[i][j] - an0[i][j] - as0[i][j]
        ap1[i][j] = ap0[i][j] + ae0[i][j] + aw0[i][j] + an0[i][j] + as0[i][j]
        bp[i][j] = s * dx * dy
    i = n
    for j in range(2, m):
        ae0[i][j] = k * dy / (dx / 2)
        aw0[i][j] = k * dy / dx
        an0[i][j] = k * dx / dy
        as0[i][j] = k * dx / dy
        ap0[i][j] = den * cv * dx * dy / dt - ae0[i][j] - aw0[i][j] - an0[i][j] - as0[i][j]
        ap1[i][j] = ap0[i][j] + ae0[i][j] + aw0[i][j] + an0[i][j] + as0[i][j]
        bp[i][j] = s * dx * dy
    for i in range(2, n):
        j = 1
        ae0[i][j]=k * dy / dx
        aw0[i][j]=k * dy / dx
        an0[i][j]=k * dx / dy
        as0[i][j]=k * dx / (dy / 2)
        ap0[i][j]=den * cv * dx * dy / dt-ae0[i][j]-aw0[i][j]-an0[i][j]-as0[i][j]
        ap1[i][j]=ap0[i][j]+ae0[i][j]+aw0[i][j]+an0[i][j]+as0[i][j]
        bp[i][j]=s * dx * dy
        j = m
        ae0[i][j]=k * dy / dx
        aw0[i][j]=k * dy / dx
        an0[i][j]=k * dx / (dy / 2)
        as0[i][j]=k * dx / dy
        ap0[i][j]=den * cv * dx * dy / dt-ae0[i][j]-aw0[i][j]-an0[i][j]-as0[i][j]
        ap1[i][j]=ap0[i][j]+ae0[i][j]+aw0[i][j]+an0[i][j]+as0[i][j]
        bp[i][j]=s * dx * dy
    i = 1
    j = 1
    ae0[i][j] = k * dy / dx
    aw0[i][j] = k * dy / (dx / 2)
    an0[i][j] = k * dx / dy
    as0[i][j] = k * dx / (dy / 2)
    ap0[i][j] = den * cv * dx * dy / dt - ae0[i][j] - aw0[i][j] - an0[i][j] - as0[i][j]
    ap1[i][j] = ap0[i][j] + ae0[i][j] + aw0[i][j] + an0[i][j] + as0[i][j]
    bp[i][j] = s * dx * dy
    i = 1
    j = m
    ae0[i][j] = k * dy / dx
    aw0[i][j] = k * dy / (dx / 2)
    an0[i][j] = k * dx / (dy / 2)
    as0[i][j] = k * dx / dy
    ap0[i][j] = den * cv * dx * dy / dt - ae0[i][j] - aw0[i][j] - an0[i][j] - as0[i][j]
    ap1[i][j] = ap0[i][j] + ae0[i][j] + aw0[i][j] + an0[i][j] + as0[i][j]
    bp[i][j] = s * dx * dy
    i = n
    j = 1
    ae0[i][j] = k * dy / (dx / 2)
    aw0[i][j] = k * dy / dx
    an0[i][j] = k * dx / dy
    as0[i][j] = k * dx / (dy / 2)
    ap0[i][j] = den * cv * dx * dy / dt - ae0[i][j] - aw0[i][j] - an0[i][j] - as0[i][j]
    ap1[i][j] = ap0[i][j] + ae0[i][j] + aw0[i][j] + an0[i][j] + as0[i][j]
    bp[i][j] = s * dx * dy
    i = n
    j = m
    ae0[i][j] = k * dy / (dx / 2)
    aw0[i][j] = k * dy / dx
    an0[i][j] = k * dx / (dy / 2)
    as0[i][j] = k * dx / dy
    ap0[i][j] = den * cv * dx * dy / dt - ae0[i][j] - aw0[i][j] - an0[i][j] - as0[i][j]
    ap1[i][j] = ap0[i][j] + ae0[i][j] + aw0[i][j] + an0[i][j] + as0[i][j]
    bp[i][j] = s * dx * dy
elif tr == 1:
    for i in range(2, n):
        for j in range(2, m):
            ae1[i][j] = k * dy / dx
            aw1[i][j] = k * dy / dx
            an1[i][j] = k * dx / dy
            as1[i][j] = k * dx / dy
            ap0[i][j] = den * cv * dx * dy / dt
            ap1[i][j] = ap0[i][j] + ae1[i][j] + aw1[i][j] + an1[i][j] + as1[i][j]
            bp[i][j] = s * dx * dy
    i = 1
    for j in range(2, m):
        ae1[i][j] = k * dy / dx
        aw1[i][j] = k * dy / (dx / 2)
        an1[i][j] = k * dx / dy
        as1[i][j] = k * dx / dy
        ap0[i][j] = den * cv * dx * dy / dt
        ap1[i][j] = ap0[i][j] + ae1[i][j] + aw1[i][j] + an1[i][j] + as1[i][j]
        bp[i][j] = s * dx * dy
    i = n
    for j in range(2, m):
        ae1[i][j] = k * dy / (dx / 2)
        aw1[i][j] = k * dy / dx
        an1[i][j] = k * dx / dy
        as1[i][j] = k * dx / dy
        ap0[i][j] = den * cv * dx * dy / dt
        ap1[i][j] = ap0[i][j] + ae1[i][j] + aw1[i][j] + an1[i][j] + as1[i][j]
        bp[i][j] = s * dx * dy
    for i in range(2, n):
        j = 1
        ae1[i][j] = k * dy / dx
        aw1[i][j] = k * dy / dx
        an1[i][j] = k * dx / dy
        as1[i][j] = k * dx / (dy / 2)
        ap0[i][j] = den * cv * dx * dy / dt
        ap1[i][j] = ap0[i][j] + ae1[i][j] + aw1[i][j] + an1[i][j] + as1[i][j]
        bp[i][j] = s * dx * dy
        j = m
        ae1[i][j] = k * dy / dx
        aw1[i][j] = k * dy / dx
        an1[i][j] = k * dx / (dy / 2)
        as1[i][j] = k * dx / dy
        ap0[i][j] = den * cv * dx * dy / dt
        ap1[i][j] = ap0[i][j] + ae1[i][j] + aw1[i][j] + an1[i][j] + as1[i][j]
        bp[i][j] = s * dx * dy
    i = 1
    j = 1
    ae1[i][j] = k * dy / dx
    aw1[i][j] = k * dy / (dx / 2)
    an1[i][j] = k * dx / dy
    as1[i][j] = k * dx / (dy / 2)
    ap0[i][j] = den * cv * dx * dy / dt
    ap1[i][j] = ap0[i][j] + ae1[i][j] + aw1[i][j] + an1[i][j] + as1[i][j]
    bp[i][j] = s * dx * dy
    i = 1
    j = m
    ae1[i][j] = k * dy / dx
    aw1[i][j] = k * dy / (dx / 2)
    an1[i][j] = k * dx / (dy / 2)
    as1[i][j] = k * dx / dy
    ap0[i][j] = den * cv * dx * dy / dt
    ap1[i][j] = ap0[i][j] + ae1[i][j] + aw1[i][j] + an1[i][j] + as1[i][j]
    bp[i][j] = s * dx * dy
    i = n
    j = 1
    ae1[i][j] = k * dy / (dx / 2)
    aw1[i][j] = k * dy / dx
    an1[i][j] = k * dx / dy
    as1[i][j] = k * dx / (dy / 2)
    ap0[i][j] = den * cv * dx * dy / dt
    ap1[i][j] = ap0[i][j] + ae1[i][j] + aw1[i][j] + an1[i][j] + as1[i][j]
    bp[i][j] = s * dx * dy
    i = n
    j = m
    ae1[i][j] = k * dy / (dx / 2)
    aw1[i][j] = k * dy / dx
    an1[i][j] = k * dx / (dy / 2)
    as1[i][j] = k * dx / dy
    ap0[i][j] = den * cv * dx * dy / dt
    ap1[i][j] = ap0[i][j] + ae1[i][j] + aw1[i][j] + an1[i][j] + as1[i][j]
    bp[i][j] = s * dx * dy


for i in range(1, n + 1):
    for j in range(1, m+1):
        T0[i][j]=3
        T[i][j]=3

if tr == 0:
    i=0
    for j in range(1, m+1):
        T0[i][j] = TL
    i= n + 1
    for j in range(1, m+1):
        T0[i][j] = TR
    j=0
    for i in range(1, n+1):
        T0[i][j] = TB
    j=m+1
    for i in range(1, n+1):
        T0[i][j] = TU

elif tr == 1:
    i=0
    for j in range(1, m+1):
        T[i][j] = TL
    i=n+1
    for j in range(1, m+1):
        T[i][j] = TR
    j=0
    for i in range(1, n+1):
        T[i][j] = TB
    j=m+1
    for i in range(1, n+1):
        T[i][j] = TU
    i=0
    for j in range(1, m+1):
        Tn0[i][j] = TL
    i=n+1
    for j in range(1, m+1):
        Tn0[i][j] = TR
    j=0
    for i in range(1, n+1):
        Tn0[i][j] = TB
    j=m+1
    for i in range(1, n+1):
        Tn0[i][j] = TU

data = ['T11', 'T21', 'T31', 'T41', 'T51', 'T12', 'T22', 'T32', 'T42','T52', 'T13', 'T23', 'T33', 'T43','T53', 'T14', 'T24', 'T34', 'T44', 'T54','T15', 'T25', 'T35', 'T45', 'T55']

a = [tstep for tstep in range(1, maxstep+1)]
b = [tstep * dt for tstep in range(1, maxstep+1)]
dataframe = pandas.DataFrame({'a':a,'time':b})
writer = pandas.ExcelWriter('C:/Users/Luogl2000/Desktop/test.xlsx')
dataframe.to_excel(writer)
writer.save()

wb = openpyxl.load_workbook('C:/Users/Luogl2000/Desktop/test.xlsx')
ws = wb['Sheet1']
for i in range(1, len(data) + 1):
    distance = data[i - 1]
    ws.cell(row = 1, column = i + 3).value = distance
wb.save('C:/Users/Luogl2000/Desktop/test.xlsx')

for tstep in range(1, maxstep + 1):
    if tr == 0:
        for j in range(1, m+1):
            for i in range(1, n+1):
                 T[i][j] = (ae0[i][j] * T0[i+1][j] + aw0[i][j] * T0[i-1][j] + an0[i][j]\
                               * T0[i][j+1] + as0[i][j] * T0[i][j-1] + ap0[i][j] * T0[i][j] + bp[i][j]) / ap1[i][j]
                 list1 = list1 + [T[i][j]]
        print(list1)
        wb = openpyxl.load_workbook('C:/Users/Luogl2000/Desktop/test.xlsx')
        ws = wb['Sheet1']
        for i in range(1, len(list1) + 1):
            distance = list1[i - 1]
            ws.cell(row = tstep + 1, column = i + 3).value = distance
        wb.save('C:/Users/Luogl2000/Desktop/test.xlsx')
        list1 = []
    elif tr == 1:
        for j in range(1, m + 1):
            for i in range(1, n + 1):
                Tn0[i][j] = T0[i][j]
        reali=maxiter
        for iter in range(1, maxiter+1):
            for j in range(1, m+1):
                for i in range(1, n+1):
                    if miter == 0:
                        T[i][j] = (ae1[i][j] * Tn0[i+1][j] + aw1[i][j] * Tn0[i-1][j] /
                                + an1[i][j] * Tn0[i][j+1] + as1[i][j] * Tn0[i][j-1] /
                                + ap0[i][j] * T0[i][j] + bp[i][j]) / ap1[i][j]
                    else:
                        T[i][j] = (ae1[i][j] * T[i+1][j] + aw1[i][j] * T[i-1][j] /
                                + an1[i][j] * T[i][j+1] + as1[i][j] * T[i][j-1] /
                                + ap0[i][j] * T0[i][j] + bp[i][j]) / ap1[i][j]
                    resi[i][j] = abs(T[i][j]-Tn0[i][j])
                    list2 = list2 + [T[i][j]]
            print(list2)
            print(len(list2))
            resimax=0
            for i in range(1, n + 1):
                for j in range(1, m + 1):
                    if resi[i][j] > resimax:
                        resimax = resi[i][j]
            if resimax < maxresi:
                reali = iter
                break
            for j in range(1, m+1):
                for i in range(1, n+1):
                    Tn0[i][j] = T[i][j]
        wb = openpyxl.load_workbook('C:/Users/Luogl2000/Desktop/test.xlsx')
        ws = wb['Sheet1']
        for i in range(1, len(list2) + 1):
            distance = list2[i - 1]
            ws.cell(row = tstep + 1, column = i + 3).value = distance
        wb.save('C:/Users/Luogl2000/Desktop/test.xlsx')
        list2 = []
        #fprintf(outputT,",%e, %i \n",resimax,reali)
    for i in range(1, n + 1):
        for j in range(1, m + 1):
            T0[i][j] = T[i][j]